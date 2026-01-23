from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from functools import wraps
import os
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-change-this-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///vapa_orders.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Database Models
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(120), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    orders = db.relationship('Order', backref='teacher', lazy=True, cascade='all, delete-orphan')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Order(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    vendor = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    items = db.relationship('OrderItem', backref='order', lazy=True, cascade='all, delete-orphan')

class OrderItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    order_id = db.Column(db.Integer, db.ForeignKey('order.id'), nullable=False)
    catalog_number = db.Column(db.String(100))
    description = db.Column(db.String(500))
    quantity = db.Column(db.Integer, default=0)
    unit_cost = db.Column(db.Float, default=0.0)
    category = db.Column(db.String(50))
    
    @property
    def total_cost(self):
        return self.quantity * self.unit_cost if self.quantity and self.unit_cost else 0

class Settings(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(100), unique=True, nullable=False)
    value = db.Column(db.String(500), nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    updated_by = db.Column(db.String(100))

class BudgetHistory(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    fiscal_year = db.Column(db.String(20), nullable=False)
    total_budget = db.Column(db.Float, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.String(100))
    notes = db.Column(db.String(500))

# Helper function to get current budget
def get_current_budget():
    setting = Settings.query.filter_by(key='total_budget').first()
    if setting:
        return float(setting.value)
    return 50128.0  # Default fallback

# Login required decorator
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Admin required decorator
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        user = User.query.get(session['user_id'])
        if not user or not user.is_admin:
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# Routes
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and user.is_active and user.check_password(password):
            session['user_id'] = user.id
            session['username'] = user.username
            session['full_name'] = user.full_name
            session['is_admin'] = user.is_admin
            
            flash(f'Welcome back, {user.full_name}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        user = User.query.get(session['user_id'])
        
        if not user.check_password(current_password):
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('change_password'))
        
        if current_password == new_password:
            flash('New password must be different from current password.', 'warning')
            return redirect(url_for('change_password'))
        
        if len(new_password) < 6:
            flash('Password must be at least 6 characters long.', 'danger')
            return redirect(url_for('change_password'))
        
        user.set_password(new_password)
        db.session.commit()
        
        flash('Password changed successfully!', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('change_password.html')

@app.route('/dashboard')
@login_required
def dashboard():
    user = User.query.get(session['user_id'])
    
    if user.is_admin:
        # Admin dashboard
        teachers = User.query.filter_by(is_admin=False).all()
        all_orders = Order.query.all()
        
        # Calculate department totals
        category_totals = {
            'Books&Supplies': 0,
            'Consultants': 0,
            'Repairs': 0,
            'SoftwareLicensing': 0,
            'ConferenceTraining': 0,
            'FieldTrips': 0
        }
        
        teacher_totals = {}
        
        for order in all_orders:
            teacher_name = order.teacher.full_name
            if teacher_name not in teacher_totals:
                teacher_totals[teacher_name] = {
                    'total': 0,
                    'categories': {cat: 0 for cat in category_totals.keys()}
                }
            
            for item in order.items:
                if item.category in category_totals:
                    category_totals[item.category] += item.total_cost
                    teacher_totals[teacher_name]['categories'][item.category] += item.total_cost
                    teacher_totals[teacher_name]['total'] += item.total_cost
        
        total_budget = get_current_budget()
        total_allocated = sum(category_totals.values())
        remaining = total_budget - total_allocated
        
        return render_template('admin_dashboard.html',
                             teachers=teachers,
                             category_totals=category_totals,
                             teacher_totals=teacher_totals,
                             total_budget=total_budget,
                             total_allocated=total_allocated,
                             remaining=remaining)
    else:
        # Teacher dashboard
        orders = Order.query.filter_by(user_id=user.id).all()
        
        # Calculate teacher's totals
        category_totals = {
            'Books&Supplies': 0,
            'Consultants': 0,
            'Repairs': 0,
            'SoftwareLicensing': 0,
            'ConferenceTraining': 0,
            'FieldTrips': 0
        }
        
        for order in orders:
            for item in order.items:
                if item.category in category_totals:
                    category_totals[item.category] += item.total_cost
        
        return render_template('teacher_dashboard.html',
                             orders=orders,
                             category_totals=category_totals)

@app.route('/admin/settings', methods=['GET', 'POST'])
@admin_required
def admin_settings():
    if request.method == 'POST':
        new_budget = request.form.get('total_budget')
        fiscal_year = request.form.get('fiscal_year')
        notes = request.form.get('notes')
        
        try:
            new_budget = float(new_budget)
            
            # Update or create budget setting
            setting = Settings.query.filter_by(key='total_budget').first()
            if setting:
                old_budget = float(setting.value)
                setting.value = str(new_budget)
                setting.updated_by = session['full_name']
                setting.updated_at = datetime.utcnow()
            else:
                setting = Settings(
                    key='total_budget',
                    value=str(new_budget),
                    updated_by=session['full_name']
                )
                db.session.add(setting)
            
            # Add to budget history
            history = BudgetHistory(
                fiscal_year=fiscal_year,
                total_budget=new_budget,
                created_by=session['full_name'],
                notes=notes
            )
            db.session.add(history)
            
            db.session.commit()
            flash(f'Budget updated to ${new_budget:,.2f} for {fiscal_year}!', 'success')
            return redirect(url_for('admin_settings'))
            
        except ValueError:
            flash('Invalid budget amount. Please enter a valid number.', 'danger')
    
    # Get current budget
    current_budget = get_current_budget()
    
    # Get budget history
    history = BudgetHistory.query.order_by(BudgetHistory.created_at.desc()).limit(10).all()
    
    # Get current fiscal year
    now = datetime.now()
    current_fiscal_year = f"{now.year}-{now.year + 1}"
    
    return render_template('admin_settings.html',
                         current_budget=current_budget,
                         history=history,
                         current_fiscal_year=current_fiscal_year)

# Add this route to app.py (after the admin_settings route, around line 215)

@app.route('/admin/orders')
@admin_required
def admin_all_orders():
    # Get filter parameters
    teacher_filter = request.args.get('teacher', '')
    category_filter = request.args.get('category', '')
    sort_by = request.args.get('sort', 'date_desc')
    
    # Get all orders
    query = Order.query
    
    # Apply teacher filter
    if teacher_filter:
        query = query.filter(Order.user_id == int(teacher_filter))
    
    # Get all orders
    all_orders = query.all()
    
    # Apply category filter and calculate totals
    filtered_orders = []
    for order in all_orders:
        order_total = sum(item.total_cost for item in order.items)
        order_categories = {}
        
        for item in order.items:
            if item.category:
                order_categories[item.category] = order_categories.get(item.category, 0) + item.total_cost
        
        # If category filter is set, only include orders with that category
        if category_filter:
            if category_filter in order_categories:
                filtered_orders.append({
                    'order': order,
                    'total': order_total,
                    'categories': order_categories,
                    'item_count': len(order.items)
                })
        else:
            filtered_orders.append({
                'order': order,
                'total': order_total,
                'categories': order_categories,
                'item_count': len(order.items)
            })
    
    # Sort orders
    if sort_by == 'date_desc':
        filtered_orders.sort(key=lambda x: x['order'].created_at, reverse=True)
    elif sort_by == 'date_asc':
        filtered_orders.sort(key=lambda x: x['order'].created_at)
    elif sort_by == 'teacher':
        filtered_orders.sort(key=lambda x: x['order'].teacher.full_name)
    elif sort_by == 'vendor':
        filtered_orders.sort(key=lambda x: x['order'].vendor)
    elif sort_by == 'total_desc':
        filtered_orders.sort(key=lambda x: x['total'], reverse=True)
    elif sort_by == 'total_asc':
        filtered_orders.sort(key=lambda x: x['total'])
    
    # Get all teachers for filter dropdown
    teachers = User.query.filter_by(is_admin=False).all()
    
    # Calculate summary statistics
    total_orders = len(filtered_orders)
    total_amount = sum(o['total'] for o in filtered_orders)
    
    categories = ['Books&Supplies', 'Consultants', 'Repairs', 'SoftwareLicensing', 'ConferenceTraining', 'FieldTrips']
    
    return render_template('admin_all_orders.html',
                         orders=filtered_orders,
                         teachers=teachers,
                         categories=categories,
                         teacher_filter=teacher_filter,
                         category_filter=category_filter,
                         sort_by=sort_by,
                         total_orders=total_orders,
                         total_amount=total_amount)


@app.route('/admin/teachers')
@admin_required
def admin_teachers():
    teachers = User.query.filter_by(is_admin=False).all()
    return render_template('admin_teachers.html', teachers=teachers)

@app.route('/admin/teachers/add', methods=['GET', 'POST'])
@admin_required
def add_teacher():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('add_teacher'))
        
        new_teacher = User(
            username=username,
            full_name=full_name,
            is_admin=False,
            is_active=True
        )
        new_teacher.set_password(password)
        
        db.session.add(new_teacher)
        db.session.commit()
        
        flash(f'Teacher {full_name} added successfully!', 'success')
        return redirect(url_for('admin_teachers'))
    
    return render_template('add_teacher.html')

# REPLACE the edit_teacher route in app.py with this updated version
# (Find the route around line 225-245 and replace it)

@app.route('/admin/teachers/<int:teacher_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_teacher(teacher_id):
    teacher = User.query.get_or_404(teacher_id)
    
    if teacher.is_admin and teacher.id == session['user_id']:
        # Special handling: user is editing themselves as admin
        # Allow everything except changing their own admin status
        pass
    
    if request.method == 'POST':
        teacher.username = request.form.get('username')
        teacher.full_name = request.form.get('full_name')
        teacher.is_active = request.form.get('is_active') == 'on'
        
        # Only update admin status if not editing yourself
        if teacher.id != session['user_id']:
            teacher.is_admin = request.form.get('is_admin') == 'on'
        
        new_password = request.form.get('password')
        if new_password:
            teacher.set_password(new_password)
        
        db.session.commit()
        
        if teacher.is_admin:
            flash(f'{teacher.full_name} updated successfully! They now have admin access.', 'success')
        else:
            flash(f'Teacher {teacher.full_name} updated successfully!', 'success')
        
        return redirect(url_for('admin_teachers'))
    
    return render_template('edit_teacher.html', teacher=teacher)


@app.route('/admin/teachers/<int:teacher_id>/delete', methods=['POST'])
@admin_required
def delete_teacher(teacher_id):
    teacher = User.query.get_or_404(teacher_id)
    
    if teacher.is_admin:
        flash('Cannot delete admin users.', 'danger')
        return redirect(url_for('admin_teachers'))
    
    db.session.delete(teacher)
    db.session.commit()
    
    flash(f'Teacher {teacher.full_name} deleted successfully.', 'success')
    return redirect(url_for('admin_teachers'))

@app.route('/orders/new', methods=['GET', 'POST'])
@login_required
def new_order():
    if request.method == 'POST':
        vendor = request.form.get('vendor')
        
        new_order = Order(
            user_id=session['user_id'],
            vendor=vendor
        )
        db.session.add(new_order)
        db.session.commit()
        
        flash(f'New order form created for {vendor}!', 'success')
        return redirect(url_for('edit_order', order_id=new_order.id))
    
    return render_template('new_order.html')

@app.route('/orders/<int:order_id>')
@login_required
def view_order(order_id):
    order = Order.query.get_or_404(order_id)
    
    # Check permissions
    user = User.query.get(session['user_id'])
    if not user.is_admin and order.user_id != user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    category_totals = {
        'Books&Supplies': 0,
        'Consultants': 0,
        'Repairs': 0,
        'SoftwareLicensing': 0,
        'ConferenceTraining': 0,
        'FieldTrips': 0
    }
    
    for item in order.items:
        if item.category in category_totals:
            category_totals[item.category] += item.total_cost
    
    return render_template('view_order.html', order=order, category_totals=category_totals)

@app.route('/orders/<int:order_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_order(order_id):
    order = Order.query.get_or_404(order_id)
    
    # Check permissions
    user = User.query.get(session['user_id'])
    if not user.is_admin and order.user_id != user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        order.vendor = request.form.get('vendor')
        
        # Update or create items
        items_data = request.form.getlist('item_id')
        
        for i, item_id in enumerate(items_data):
            catalog = request.form.get(f'catalog_{i}')
            description = request.form.get(f'description_{i}')
            quantity = request.form.get(f'quantity_{i}')
            unit_cost = request.form.get(f'unit_cost_{i}')
            category = request.form.get(f'category_{i}')
            
            # Skip empty rows
            if not any([catalog, description, quantity, unit_cost]):
                continue
            
            if item_id and item_id.isdigit():
                # Update existing item
                item = OrderItem.query.get(int(item_id))
                if item and item.order_id == order.id:
                    item.catalog_number = catalog
                    item.description = description
                    item.quantity = int(quantity) if quantity else 0
                    item.unit_cost = float(unit_cost) if unit_cost else 0.0
                    item.category = category
            else:
                # Create new item
                new_item = OrderItem(
                    order_id=order.id,
                    catalog_number=catalog,
                    description=description,
                    quantity=int(quantity) if quantity else 0,
                    unit_cost=float(unit_cost) if unit_cost else 0.0,
                    category=category
                )
                db.session.add(new_item)
        
        db.session.commit()
        flash('Order updated successfully!', 'success')
        return redirect(url_for('view_order', order_id=order.id))
    
    return render_template('edit_order.html', order=order)

@app.route('/orders/<int:order_id>/delete', methods=['POST'])
@login_required
def delete_order(order_id):
    order = Order.query.get_or_404(order_id)
    
    # Check permissions
    user = User.query.get(session['user_id'])
    if not user.is_admin and order.user_id != user.id:
        flash('Access denied.', 'danger')
        return redirect(url_for('dashboard'))
    
    db.session.delete(order)
    db.session.commit()
    
    flash('Order deleted successfully.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/export/excel')
@admin_required
def export_excel():
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Create summary sheet
    summary = wb.create_sheet('Summary')
    summary['A1'] = 'VAPA Department Budget Summary'
    summary['A1'].font = Font(bold=True, size=14)
    
    summary['A3'] = 'Category'
    summary['B3'] = 'Total'
    summary['A3'].fill = header_fill
    summary['B3'].fill = header_fill
    summary['A3'].font = header_font
    summary['B3'].font = header_font
    
    categories = ['Books&Supplies', 'Consultants', 'Repairs', 'SoftwareLicensing', 'ConferenceTraining', 'FieldTrips']
    category_totals = {cat: 0 for cat in categories}
    
    all_orders = Order.query.all()
    for order in all_orders:
        for item in order.items:
            if item.category in category_totals:
                category_totals[item.category] += item.total_cost
    
    row = 4
    for cat in categories:
        summary[f'A{row}'] = cat
        summary[f'B{row}'] = category_totals[cat]
        summary[f'B{row}'].number_format = '$#,##0.00'
        row += 1
    
    # Create sheet for each teacher
    teachers = User.query.filter_by(is_admin=False).all()
    for teacher in teachers:
        sheet = wb.create_sheet(teacher.full_name)
        
        sheet['A1'] = f'{teacher.full_name} - Orders'
        sheet['A1'].font = Font(bold=True, size=14)
        
        row = 3
        for order in teacher.orders:
            sheet[f'A{row}'] = f'Vendor: {order.vendor}'
            sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            headers = ['Catalog #', 'Description', 'Quantity', 'Unit Cost', 'Total Cost', 'Category']
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
            
            row += 1
            
            for item in order.items:
                sheet.cell(row=row, column=1).value = item.catalog_number
                sheet.cell(row=row, column=2).value = item.description
                sheet.cell(row=row, column=3).value = item.quantity
                sheet.cell(row=row, column=4).value = item.unit_cost
                sheet.cell(row=row, column=4).number_format = '$#,##0.00'
                sheet.cell(row=row, column=5).value = item.total_cost
                sheet.cell(row=row, column=5).number_format = '$#,##0.00'
                sheet.cell(row=row, column=6).value = item.category
                row += 1
            
            row += 2
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'VAPA_Orders_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

@app.route('/api/order/<int:order_id>/item/<int:item_id>/delete', methods=['POST'])
@login_required
def delete_item(order_id, item_id):
    item = OrderItem.query.get_or_404(item_id)
    order = item.order
    
    # Check permissions
    user = User.query.get(session['user_id'])
    if not user.is_admin and order.user_id != user.id:
        return jsonify({'success': False, 'message': 'Access denied'}), 403
    
    # Delete the item
    db.session.delete(item)
    db.session.commit()
    
    # Check if order now has no items - if so, delete the entire order
    remaining_items = OrderItem.query.filter_by(order_id=order.id).count()
    if remaining_items == 0:
        db.session.delete(order)
        db.session.commit()
        return jsonify({'success': True, 'order_deleted': True})
    
    return jsonify({'success': True, 'order_deleted': False})

# Initialize database
def init_db():
    with app.app_context():
        db.create_all()
        
        # Create default admin if not exists
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                full_name='Administrator',
                is_admin=True,
                is_active=True
            )
            admin.set_password('admin123')
            db.session.add(admin)
            db.session.commit()
            print('Default admin created: username=admin, password=admin123')
            print('IMPORTANT: Change the admin password after first login!')
        
        # Create default budget setting if not exists
        budget_setting = Settings.query.filter_by(key='total_budget').first()
        if not budget_setting:
            budget_setting = Settings(
                key='total_budget',
                value='50128',
                updated_by='System'
            )
            db.session.add(budget_setting)
            db.session.commit()
            print('Default budget created: $50,128')

if __name__ == '__main__':
    init_db()
    app.run(debug=True)
